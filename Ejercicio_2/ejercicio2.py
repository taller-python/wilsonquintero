"""Modulo para leer el archivo de excel operaciones.xlsx con sus hojas
y realizar operaciones de suma, resta multiplicacion
y división; y escribir los resultados en un archivo txt"""

import openpyxl

WB = openpyxl.load_workbook(filename='./Ejercicio_2/operaciones.xlsx', read_only=True)


# Hoja Suma


HOJA = WB['SUMA']

RES = open('./Ejercicio_2/RESULTADO.txt', 'w')
RES.write("HOJA SUMA...\n")
for n1, n2 in HOJA.iter_rows(min_row=2):

                try: 

                        SUMA =  int(n1.value + n2.value)
                except:
                        SUMA = 'Error'
                print(str(SUMA))
                RES.write(str(SUMA)+"\n")

        
# Hoja Resta
HOJA = WB['RESTA']

RES.write("HOJA RESTA...\n")
for n1, n2 in HOJA.iter_rows(min_row=2):

                try: 

                        RESTA =  int(n1.value - n2.value)
                except:
                        RESTA = 'Error'
                print(str(RESTA))
                RES.write(str(RESTA)+"\n")

# Hoja MULTIPLICACION
HOJA = WB['MULTIPLICACIÓN']

RES.write("HOJA MULTIPLICACIÓN...\n")
for n1, n2 in HOJA.iter_rows(min_row=2):

                try: 

                        MULT =  int(n1.value * n2.value)
                except:
                        MULT = 'Error'
                print(str(MULT))
                RES.write(str(MULT)+"\n")


# Hoja DIVISIÓN
HOJA = WB['DIVISIÓN']

RES.write("HOJA DIVKISIÓN...\n")
for n1, n2 in HOJA.iter_rows(min_row=2):

                try: 

                        DIV =  int(n1.value / n2.value)
                except ZeroDivisionError:
                        DIV = 'Div. por 0 no Existe'
                except:
                        DIV = 'Error'
                print(str(DIV))
                RES.write(str(DIV)+"\n")
RES.close()
